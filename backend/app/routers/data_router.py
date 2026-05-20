import io
import json
import shutil
import csv
import os
import tempfile
import zipfile
from pathlib import Path
from typing import Optional, List, Tuple
from urllib.parse import quote

from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Depends
from fastapi.responses import StreamingResponse

from ..auth import get_user_db, require_permission
from ..sql_utils import build_where, build_search_clause, build_where_inline, build_search_clause_inline
from ..storage import get_storage

router = APIRouter(prefix="/api", tags=["data"])


def _sanitize_table_name(filename: str) -> str:
    return Path(filename).stem.replace("-", "_").replace(" ", "_").lower()


def _detect_and_decode(raw: bytes) -> str:
    """Detect encoding of raw bytes and decode to str. Handles GBK/GB2312/GB18030 etc."""
    import chardet
    # Check for UTF-8 BOM
    if raw.startswith(b"\xef\xbb\xbf"):
        return raw[3:].decode("utf-8", errors="replace")
    # Use chardet to detect encoding first
    result = chardet.detect(raw)
    encoding = result.get("encoding") or "utf-8"
    confidence = result.get("confidence") or 0
    # chardet sometimes returns 'GB2312' but the file may use the full GB18030 range
    if encoding.lower() in ("gb2312", "gbk", "gb18030", "iso-8859-1", "windows-1252"):
        # For CJK files, ISO-8859-1/Windows-1252 is almost always a misdetection
        # Try GB18030 first for Chinese content
        try:
            return raw.decode("gb18030")
        except (UnicodeDecodeError, LookupError):
            pass
    # If chardet is confident about UTF-8, use it
    if encoding.lower() == "utf-8" and confidence > 0.8:
        return raw.decode("utf-8", errors="replace")
    # Try detected encoding
    try:
        return raw.decode(encoding, errors="replace")
    except (UnicodeDecodeError, LookupError):
        return raw.decode("utf-8", errors="replace")


def _preprocess_csv(file_path: Path, comment_char: str = "#"):
    """Remove lines starting with comment_char from a CSV file (in-place)."""
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
    filtered = [line for line in lines if not line.lstrip().startswith(comment_char)]
    with open(file_path, "w", encoding="utf-8") as f:
        f.writelines(filtered)


def _read_file_as_csv_path(file: UploadFile, user_id: str) -> Path:
    """Save uploaded file to local storage, converting xlsx to csv if needed.
    For CSV files, detect encoding and re-encode as UTF-8."""
    csv_filename, csv_bytes = _normalize_to_csv_bytes(file.filename, file.file.read())
    storage = get_storage()
    storage.save(user_id, csv_filename, csv_bytes)
    return storage.get_local_path(user_id, csv_filename)


def _save_local_file_as_csv_path(local_path: Path, save_as_filename: str, user_id: str) -> Path:
    """Same as _read_file_as_csv_path but reads from a local extracted file (e.g. from a zip).
    `save_as_filename` is the name used for storage — pass a path-unique name to avoid
    collisions when multiple files in a zip share the same basename."""
    csv_filename, csv_bytes = _normalize_to_csv_bytes(save_as_filename, local_path.read_bytes())
    storage = get_storage()
    storage.save(user_id, csv_filename, csv_bytes)
    return storage.get_local_path(user_id, csv_filename)


def _normalize_to_csv_bytes(filename: str, raw: bytes) -> Tuple[str, bytes]:
    """Convert xlsx bytes to UTF-8 CSV bytes; for CSV input, re-encode to UTF-8.
    Returns (final_filename, utf8_bytes)."""
    if filename.lower().endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        csv_filename = Path(filename).stem + ".csv"
        output = io.StringIO()
        writer = csv.writer(output)

        sheets = wb.sheetnames
        if len(sheets) > 1:
            # Multi-sheet: merge all sheets with the same structure
            header = None
            for sheet_name in sheets:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                sheet_header = rows[0]
                if header is None:
                    header = sheet_header
                    writer.writerow(header)
                    for row in rows[1:]:
                        writer.writerow(row)
                elif tuple(sheet_header) == tuple(header):
                    # Same structure, skip header and append data
                    for row in rows[1:]:
                        writer.writerow(row)
                # Different structure sheets are silently skipped
        else:
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)

        wb.close()
        return csv_filename, output.getvalue().encode("utf-8")

    # CSV: detect encoding and convert to UTF-8
    text = _detect_and_decode(raw)
    return filename, text.encode("utf-8")


def _duckdb_read_csv(conn, save_path: Path, table_name: str):
    """Create a table from CSV using DuckDB read_csv_auto with proper path escaping."""
    path_str = str(save_path.resolve()).replace("'", "''")
    conn.execute(f"""CREATE TABLE "{table_name}" AS SELECT * FROM read_csv_auto('{path_str}', header=true)""")


# ===== ZIP HELPERS =====
_ZIP_ALLOWED_EXTS = (".csv", ".xlsx")
_ZIP_SKIP_NAMES = {".DS_Store"}
_ZIP_SKIP_PREFIXES = ("__MACOSX/",)


def _safe_extract_zip(zip_bytes: bytes, dest_dir: Path) -> List[Tuple[str, Path]]:
    """Extract a zip into dest_dir with zip-slip protection and content validation.
    Returns [(relative_path_in_zip, absolute_extracted_path), ...] sorted by relative_path.
    Raises HTTPException(400) if any file inside is not csv/xlsx (excluding macOS metadata)."""
    try:
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    except zipfile.BadZipFile:
        raise HTTPException(400, "无法解析 ZIP 文件，可能已损坏")

    bad: List[str] = []
    items: List[str] = []
    with zf:
        for name in zf.namelist():
            if name.endswith("/"):
                continue
            base = Path(name).name
            if base in _ZIP_SKIP_NAMES or any(name.startswith(p) for p in _ZIP_SKIP_PREFIXES):
                continue
            if not name.lower().endswith(_ZIP_ALLOWED_EXTS):
                bad.append(name)
                continue
            items.append(name)

        if bad:
            preview = ", ".join(bad[:10])
            suffix = f" 等共 {len(bad)} 个" if len(bad) > 10 else ""
            raise HTTPException(400, f"ZIP 中包含不支持的文件（仅支持 csv/xlsx）：{preview}{suffix}")

        if not items:
            raise HTTPException(400, "ZIP 中没有可导入的 csv/xlsx 文件")

        dest_resolved = dest_dir.resolve()
        results: List[Tuple[str, Path]] = []
        for name in items:
            target = (dest_dir / name).resolve()
            try:
                target.relative_to(dest_resolved)
            except ValueError:
                raise HTTPException(400, f"ZIP 包含非法路径：{name}")
            target.parent.mkdir(parents=True, exist_ok=True)
            with zf.open(name) as src, open(target, "wb") as dst:
                shutil.copyfileobj(src, dst)
            results.append((name, target))

    results.sort(key=lambda x: x[0])
    return results


def _zip_default_table_name(relative_path: str) -> str:
    """Convert a zip-internal path to a default table name. dir/sub/a.csv → dir_sub_a."""
    p = Path(relative_path)
    parts = [seg for seg in (*p.parent.parts, p.stem) if seg not in (".", "")]
    return _sanitize_table_name("_".join(parts))


def _zip_storage_filename(relative_path: str) -> str:
    """Build a path-unique filename for storage so that dir1/a.csv and dir2/a.csv don't collide."""
    p = Path(relative_path)
    parts = [seg for seg in (*p.parent.parts, p.stem) if seg not in (".", "")]
    safe_stem = "_".join(parts).replace("-", "_").replace(" ", "_").lower()
    return f"{safe_stem}{p.suffix.lower()}"


def _validate_upload_ext(filename: str) -> Optional[str]:
    """检查文件扩展名是否支持。返回 None 表示通过，否则返回错误描述。
    注意：扩展名校验统一转小写，避免 .XLSX / .Xlsx 之类被误拒。"""
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".xlsx"):
        return None
    if name.endswith(".xls"):
        return "不支持旧版 .xls 格式，请在 Excel 中另存为 .xlsx 后再上传"
    return "只支持 CSV 和 XLSX 文件"


# ===== UPLOAD (single file) =====
@router.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    mode: str = Form("replace"),
    table_name: Optional[str] = Form(None),
    skip_comments: str = Form("false"),
    user=Depends(require_permission("upload")),
):
    err = _validate_upload_ext(file.filename)
    if err:
        raise HTTPException(400, err)
    if mode not in ("replace", "append"):
        raise HTTPException(400, "mode 参数只支持 replace 或 append")

    user_id = user["sub"]
    save_path = _read_file_as_csv_path(file, user_id)
    if skip_comments.lower() in ("true", "1", "yes"):
        _preprocess_csv(save_path)
    table_name = table_name or _sanitize_table_name(file.filename)
    conn = get_user_db(user_id)

    try:
        if mode == "append":
            # Check if table exists
            existing_tables = [t[0] for t in conn.execute("SHOW TABLES").fetchall()]
            if table_name not in existing_tables:
                raise HTTPException(400, f"表 '{table_name}' 不存在，无法追加。请先用 replace 模式创建表。")
            # Use staging table to validate columns
            staging = f"_staging_{table_name}"
            conn.execute(f'DROP TABLE IF EXISTS "{staging}"')
            _duckdb_read_csv(conn, save_path, staging)
            # Verify column names match
            existing_cols = [c[0] for c in conn.execute(f'DESCRIBE "{table_name}"').fetchall()]
            staging_cols = {c[0] for c in conn.execute(f'DESCRIBE "{staging}"').fetchall()}
            if set(existing_cols) != staging_cols:
                conn.execute(f'DROP TABLE "{staging}"')
                raise HTTPException(400, f"列结构不匹配。现有列: {set(existing_cols)}, 上传列: {staging_cols}")
            # Insert with column order matching target table
            col_list = ", ".join(f'"{c}"' for c in existing_cols)
            conn.execute(f'INSERT INTO "{table_name}" ({col_list}) SELECT {col_list} FROM "{staging}"')
            conn.execute(f'DROP TABLE "{staging}"')
        else:
            conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
            _duckdb_read_csv(conn, save_path, table_name)

        cols = conn.execute(f'DESCRIBE "{table_name}"').fetchall()
        row_count = conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
        return {
            "success": True, "table": table_name, "filename": file.filename,
            "row_count": row_count, "columns": [{"name": c[0], "type": c[1]} for c in cols],
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


# ===== BATCH UPLOAD =====
@router.post("/upload/batch")
async def upload_batch(
    files: List[UploadFile] = File(...),
    table_name: Optional[str] = Form(None),
    mode: str = Form("replace"),
    user=Depends(require_permission("upload")),
):
    if not files:
        raise HTTPException(400, "请至少上传一个文件")

    user_id = user["sub"]
    results = []

    for idx, file in enumerate(files):
        ext_err = _validate_upload_ext(file.filename)
        if ext_err:
            results.append({"filename": file.filename, "success": False, "error": ext_err})
            continue

        save_path = _read_file_as_csv_path(file, user_id)
        tname = table_name or _sanitize_table_name(file.filename)
        conn = get_user_db(user_id)

        try:
            current_mode = mode if idx == 0 else "append"
            if current_mode == "append":
                existing_tables = [t[0] for t in conn.execute("SHOW TABLES").fetchall()]
                if tname not in existing_tables:
                    current_mode = "replace"

            if current_mode == "replace":
                conn.execute(f'DROP TABLE IF EXISTS "{tname}"')
                _duckdb_read_csv(conn, save_path, tname)
            else:
                staging = f"_staging_{tname}"
                conn.execute(f'DROP TABLE IF EXISTS "{staging}"')
                _duckdb_read_csv(conn, save_path, staging)
                existing_cols = [c[0] for c in conn.execute(f'DESCRIBE "{tname}"').fetchall()]
                staging_cols = {c[0] for c in conn.execute(f'DESCRIBE "{staging}"').fetchall()}
                if set(existing_cols) != staging_cols:
                    conn.execute(f'DROP TABLE "{staging}"')
                    results.append({"filename": file.filename, "success": False, "error": f"列结构不匹配"})
                    continue
                col_list = ", ".join(f'"{c}"' for c in existing_cols)
                conn.execute(f'INSERT INTO "{tname}" ({col_list}) SELECT {col_list} FROM "{staging}"')
                conn.execute(f'DROP TABLE "{staging}"')

            row_count = conn.execute(f'SELECT COUNT(*) FROM "{tname}"').fetchone()[0]
            results.append({"filename": file.filename, "success": True, "table": tname, "row_count": row_count})
        except Exception as e:
            results.append({"filename": file.filename, "success": False, "error": str(e)})

    return {"results": results}


# ===== XLSX MULTI-SHEET PREVIEW / IMPORT =====
def _xlsx_extract_groups(raw: bytes):
    """打开 xlsx，按表头分组 sheet。返回 (workbook, groups)。
    groups 形如 [{group_id, sheet_names, header, columns, row_count, default_table_name}]，
    同表头的 sheet 会合并到同一组；row_count 是该组所有 sheet 数据行（不含表头）的合计。"""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"无法解析 xlsx 文件：{e}")

    groups: list = []  # 保留出现顺序
    header_index: dict = {}  # header_tuple -> idx in groups
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # max_row 在 read_only 模式下偶尔不准；用 iter_rows 数一遍数据行（不含表头）
        row_iter = ws.iter_rows(values_only=True)
        first_row = next(row_iter, None)
        if not first_row:
            continue
        data_rows = sum(1 for _ in row_iter)
        header = tuple("" if c is None else str(c) for c in first_row)
        if header in header_index:
            g = groups[header_index[header]]
            g["sheet_names"].append(sheet_name)
            g["row_count"] += data_rows
        else:
            header_index[header] = len(groups)
            header_list = list(header)
            groups.append({
                "group_id": len(groups),
                "sheet_names": [sheet_name],
                "header": header_list,            # 前端模板用的字段名
                "columns": header_list,           # 同义字段，兼容老调用方
                "row_count": data_rows,
                "default_table_name": _sanitize_table_name(sheet_name),
            })
    return wb, groups


@router.post("/upload/xlsx-preview")
async def upload_xlsx_preview(
    file: UploadFile = File(...),
    user=Depends(require_permission("upload")),
):
    """多 Sheet xlsx 预览：按表头分组返回，让用户在前端勾选哪些组要导入。
    本接口不持久化文件，前端在 xlsx-import 时会再次上传同一文件。"""
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(400, "只支持 .xlsx 文件")
    raw = await file.read()
    wb, groups = _xlsx_extract_groups(raw)
    wb.close()
    if not groups:
        raise HTTPException(400, "Excel 文件中没有可用的 Sheet（所有 Sheet 都为空）")
    return {"filename": file.filename, "groups": groups}


@router.post("/upload/xlsx-import")
async def upload_xlsx_import(
    file: UploadFile = File(...),
    plan: str = Form(...),
    mode: str = Form("replace"),
    skip_comments: str = Form("false"),
    user=Depends(require_permission("upload")),
):
    """多 Sheet xlsx 导入：按 plan 中的分组，把同表头的多个 Sheet 合并成一个表入库。
    plan JSON 形如 [{"group_id", "table_name", "sheet_names": [...], "include": bool}]。"""
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(400, "只支持 .xlsx 文件")
    try:
        plan_obj = json.loads(plan)
    except json.JSONDecodeError:
        raise HTTPException(400, "plan 不是合法 JSON")
    if not isinstance(plan_obj, list) or not plan_obj:
        raise HTTPException(400, "plan 不能为空")

    raw = await file.read()
    user_id = user["sub"]
    conn = get_user_db(user_id)
    storage = get_storage()

    wb, _existing_groups = _xlsx_extract_groups(raw)
    available_sheets = set(wb.sheetnames)
    results = []

    try:
        for group in plan_obj:
            if not group.get("include", True):
                continue
            table_name = (group.get("table_name") or "").strip()
            if not table_name:
                results.append({"success": False, "error": "缺少表名"})
                continue
            sheet_names = group.get("sheet_names") or []
            sheet_names = [s for s in sheet_names if s in available_sheets]
            if not sheet_names:
                results.append({"success": False, "error": f"分组 {table_name} 中没有有效的 Sheet", "table": table_name})
                continue

            try:
                # 合并同分组下的所有 sheet 为一个 CSV
                output = io.StringIO()
                writer = csv.writer(output)
                header_written = False
                for sname in sheet_names:
                    ws = wb[sname]
                    rows_iter = ws.iter_rows(values_only=True)
                    first = next(rows_iter, None)
                    if first is None:
                        continue
                    if not header_written:
                        writer.writerow(first)
                        header_written = True
                    for row in rows_iter:
                        writer.writerow(row)
                if not header_written:
                    results.append({"success": False, "error": "分组内所有 Sheet 都没有数据", "table": table_name})
                    continue

                csv_bytes = output.getvalue().encode("utf-8")
                csv_filename = f"{table_name}.csv"
                storage.save(user_id, csv_filename, csv_bytes)
                save_path = storage.get_local_path(user_id, csv_filename)
                if skip_comments.lower() in ("true", "1", "yes"):
                    _preprocess_csv(save_path)

                conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
                _duckdb_read_csv(conn, save_path, table_name)
                row_count = conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
                cols = conn.execute(f'DESCRIBE "{table_name}"').fetchall()
                results.append({
                    "success": True,
                    "table": table_name,
                    "row_count": row_count,
                    "columns": [{"name": c[0], "type": c[1]} for c in cols],
                    "merged_sheets": sheet_names,
                })
            except Exception as e:
                results.append({"success": False, "error": str(e), "table": table_name})
    finally:
        wb.close()

    return {"results": results}


# ===== ZIP UPLOAD =====
@router.post("/upload/zip-preview")
async def upload_zip_preview(
    file: UploadFile = File(...),
    user=Depends(require_permission("upload")),
):
    """Step 1 of zip upload: extract & validate, return file list for the user to confirm.
    The zip is NOT persisted — frontend will resend it in /upload/zip-import after confirmation."""
    if not file.filename.lower().endswith(".zip"):
        raise HTTPException(400, "只支持 .zip 文件")
    raw = await file.read()
    with tempfile.TemporaryDirectory() as td:
        items = _safe_extract_zip(raw, Path(td))
        return {
            "filename": file.filename,
            "files": [
                {
                    "relative_path": rel,
                    "default_table_name": _zip_default_table_name(rel),
                    "size": local.stat().st_size,
                }
                for rel, local in items
            ],
        }


@router.post("/upload/zip-import")
async def upload_zip_import(
    file: UploadFile = File(...),
    plan: str = Form(...),
    user=Depends(require_permission("upload")),
):
    """Step 2 of zip upload: re-extract the zip and import each file according to plan.
    plan JSON shape: {"skip_comments": bool, "items": [{"relative_path", "table_name"}]}.
    Always uses replace mode (zip-mode does not support append)."""
    if not file.filename.lower().endswith(".zip"):
        raise HTTPException(400, "只支持 .zip 文件")
    try:
        plan_obj = json.loads(plan)
    except json.JSONDecodeError:
        raise HTTPException(400, "plan 不是合法 JSON")

    skip_comments = bool(plan_obj.get("skip_comments", False))
    items_plan = plan_obj.get("items") or []
    if not items_plan:
        raise HTTPException(400, "plan.items 不能为空")

    rel_to_table = {it["relative_path"]: (it.get("table_name") or "").strip() for it in items_plan}
    raw = await file.read()
    user_id = user["sub"]
    conn = get_user_db(user_id)
    results = []

    with tempfile.TemporaryDirectory() as td:
        extracted = _safe_extract_zip(raw, Path(td))
        extracted_map = {rel: local for rel, local in extracted}
        missing = [rel for rel in rel_to_table if rel not in extracted_map]
        if missing:
            raise HTTPException(400, f"以下文件在 ZIP 中不存在：{missing[:5]}")

        for rel, table_name in rel_to_table.items():
            local_path = extracted_map[rel]
            try:
                save_path = _save_local_file_as_csv_path(
                    local_path, _zip_storage_filename(rel), user_id
                )
                if skip_comments:
                    _preprocess_csv(save_path)
                tname = table_name or _zip_default_table_name(rel)
                conn.execute(f'DROP TABLE IF EXISTS "{tname}"')
                _duckdb_read_csv(conn, save_path, tname)
                row_count = conn.execute(f'SELECT COUNT(*) FROM "{tname}"').fetchone()[0]
                cols = conn.execute(f'DESCRIBE "{tname}"').fetchall()
                results.append({
                    "relative_path": rel, "success": True, "table": tname,
                    "row_count": row_count,
                    "columns": [{"name": c[0], "type": c[1]} for c in cols],
                })
            except Exception as e:
                results.append({"relative_path": rel, "success": False, "error": str(e)})

    return {"results": results}


# ===== LIST TABLES =====
@router.get("/tables")
async def list_tables(user=Depends(require_permission("read"))):
    conn = get_user_db(user["sub"])
    try:
        result = []
        for (t,) in conn.execute("SHOW TABLES").fetchall():
            if t.startswith("_staging_"):
                continue
            count = conn.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0]
            cols = conn.execute(f'DESCRIBE "{t}"').fetchall()
            result.append({
                "name": t, "row_count": count,
                "columns": [{"name": c[0], "type": c[1]} for c in cols],
            })
        return {"tables": result}
    except Exception as e:
        raise HTTPException(500, str(e))


# ===== GET TABLE DATA =====
@router.get("/table/{table_name}/data")
async def get_table_data(
    table_name: str,
    page: int = 1,
    page_size: int = 50,
    sort_col: Optional[str] = None,
    sort_dir: str = "asc",
    filters: Optional[str] = None,
    search: Optional[str] = None,
    skip_count: bool = False,
    user=Depends(require_permission("read")),
):
    conn = get_user_db(user["sub"])
    try:
        params = []
        # 获取列类型信息，用于在时间列上做正确的 CAST
        cols = conn.execute(f'DESCRIBE "{table_name}"').fetchall()
        col_list = [{"name": c[0], "type": c[1]} for c in cols]
        where_clause, params = build_where(json.loads(filters) if filters else [], params, col_list)

        # Build search clause
        if search:
            search_clause, params = build_search_clause(col_list, search, params)
            if search_clause:
                if where_clause:
                    where_clause = where_clause + " AND " + search_clause
                else:
                    where_clause = "WHERE " + search_clause

        order_clause = f'ORDER BY "{sort_col}" {"DESC" if sort_dir.lower() == "desc" else "ASC"}' if sort_col else ""
        offset = (page - 1) * page_size

        # 翻页 / 调整 page size 时筛选条件没变，前端会传 skip_count=true 让我们跳过 COUNT，
        # 这是大表搜索 / 复杂筛选时的主要省时点（COUNT 必须全表扫描，无法用 LIMIT 短路）。
        total = None
        if not skip_count:
            total = conn.execute(f'SELECT COUNT(*) FROM "{table_name}" {where_clause}', params).fetchone()[0]
        rows_raw = conn.execute(
            f'SELECT * FROM "{table_name}" {where_clause} {order_clause} LIMIT {page_size} OFFSET {offset}', params
        ).fetchall()
        col_names = [c["name"] for c in col_list]
        rows = [dict(zip(col_names, row)) for row in rows_raw]
        for row in rows:
            for k, v in row.items():
                if not isinstance(v, (str, int, float, bool, type(None))):
                    row[k] = str(v)
        return {
            "columns": col_list,
            "rows": rows,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (max(1, (total + page_size - 1) // page_size) if total is not None else None),
            "skipped_count": skip_count,
        }
    except Exception as e:
        raise HTTPException(500, str(e))


# ===== EXPORT CSV (streaming, low memory) =====
@router.get("/table/{table_name}/export")
async def export_csv(
    table_name: str,
    sort_col: Optional[str] = None,
    sort_dir: str = "asc",
    filters: Optional[str] = None,
    search: Optional[str] = None,
    export_all: bool = False,
    user=Depends(require_permission("read")),
):
    """Stream table data as CSV.

    优化点：
    - 不再使用 DuckDB COPY 写临时文件，减少磁盘 IO
    - 直接 SELECT *，用 csv.writer 分块 fetchmany + 流式输出，降低内存占用
    - 仍然复用 build_where_inline / build_search_clause_inline，保证和列表页筛选逻辑一致
    """
    conn = get_user_db(user["sub"])
    try:
        # 提前拿一次 DESCRIBE，给 filter 和 search 共享列类型信息
        cols_desc = conn.execute(f'DESCRIBE "{table_name}"').fetchall()
        col_list = [{"name": c[0], "type": c[1]} for c in cols_desc]
        # 构造 WHERE 子句
        if export_all:
            where_clause = ""
        else:
            where_clause = build_where_inline(json.loads(filters) if filters else [], col_list)
            if search:
                search_clause = build_search_clause_inline(col_list, search)
                if search_clause:
                    if where_clause:
                        where_clause = where_clause + " AND " + search_clause
                    else:
                        where_clause = "WHERE " + search_clause

        order_clause = (
            f'ORDER BY "{sort_col}" {"DESC" if sort_dir.lower() == "desc" else "ASC"}'
            if sort_col
            else ""
        )
        sql = f'SELECT * FROM "{table_name}" {where_clause} {order_clause}'

        # 执行查询，保留 cursor 用于流式 fetch
        cur = conn.execute(sql)
        cols = [d[0] for d in cur.description]

        def generate():
            # 先输出 UTF-8 BOM，兼容 Excel
            yield b"\xef\xbb\xbf"

            buffer = io.StringIO()
            writer = csv.writer(buffer)

            # 写表头
            writer.writerow(cols)
            chunk = buffer.getvalue().encode("utf-8")
            if chunk:
                yield chunk
            buffer.seek(0)
            buffer.truncate(0)

            # 分批拉取数据，避免一次性加载过多行
            while True:
                rows = cur.fetchmany(10000)
                if not rows:
                    break
                for row in rows:
                    writer.writerow(row)
                chunk = buffer.getvalue().encode("utf-8")
                if chunk:
                    yield chunk
                buffer.seek(0)
                buffer.truncate(0)

        encoded_filename = quote(f"{table_name}_export.csv")
        return StreamingResponse(
            generate(),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": (
                    f"attachment; filename=\"export.csv\"; "
                    f"filename*=UTF-8''{encoded_filename}"
                ),
            },
        )
    except Exception as e:
        raise HTTPException(500, str(e))


# ===== GROUP STATS =====
_NUMERIC_TYPE_KEYWORDS = ("INT", "FLOAT", "DOUBLE", "DECIMAL", "NUMERIC", "REAL", "BIGINT", "HUGEINT", "TINYINT", "SMALLINT")
_AGG_OPS = {"COUNT", "SUM", "AVG", "MIN", "MAX"}


def _is_numeric_type(col_type: str) -> bool:
    t = (col_type or "").upper()
    return any(k in t for k in _NUMERIC_TYPE_KEYWORDS)


@router.get("/table/{table_name}/group-stats")
async def group_stats(
    table_name: str,
    group_by: str,
    aggs: Optional[str] = None,
    filters: Optional[str] = None,
    search: Optional[str] = None,
    sort_col: Optional[str] = None,
    sort_dir: str = "desc",
    page: int = 1,
    page_size: int = 50,
    skip_count: bool = False,
    user=Depends(require_permission("read")),
):
    """对一到两列做 GROUP BY 聚合统计。

    group_by: 单个列名 或 JSON 列表（如 ["col1","col2"]）；最多 2 列。
    aggs JSON: [{"op": "COUNT"|"SUM"|"AVG"|"MIN"|"MAX", "col": "目标列(COUNT 可省略表示 *)", "alias": "结果列名(可选)"}]
    复用列表页的 filters / search，使分组结果与列表展示保持一致。
    """
    conn = get_user_db(user["sub"])
    try:
        cols = conn.execute(f'DESCRIBE "{table_name}"').fetchall()
        col_list = [{"name": c[0], "type": c[1]} for c in cols]
        col_names = {c["name"] for c in col_list}
        col_type_map = {c["name"]: (c["type"] or "").upper() for c in col_list}

        # group_by 同时兼容单列名和 JSON 列表
        gb_raw = (group_by or "").strip()
        if gb_raw.startswith("["):
            try:
                group_by_list = json.loads(gb_raw)
            except json.JSONDecodeError:
                raise HTTPException(400, "group_by 不是合法的 JSON 列表")
        else:
            group_by_list = [gb_raw] if gb_raw else []
        group_by_list = [str(c).strip() for c in group_by_list if str(c).strip()]
        if not group_by_list:
            raise HTTPException(400, "请至少选择一个分组字段")
        if len(group_by_list) > 2:
            raise HTTPException(400, "最多支持 2 个分组字段")
        if len(set(group_by_list)) != len(group_by_list):
            raise HTTPException(400, "分组字段不能重复")
        for gb in group_by_list:
            if gb not in col_names:
                raise HTTPException(400, f"分组字段 '{gb}' 不存在")

        # 默认聚合：COUNT(*)
        if aggs:
            try:
                aggs_list = json.loads(aggs)
            except json.JSONDecodeError:
                raise HTTPException(400, "aggs 参数不是合法 JSON")
        else:
            aggs_list = [{"op": "COUNT", "col": "*", "alias": "count"}]

        if not isinstance(aggs_list, list) or not aggs_list:
            raise HTTPException(400, "aggs 至少要有一个聚合项")

        # 构造 SELECT 列表 + 别名校验
        select_parts = [f'"{gb}" AS "{gb}"' for gb in group_by_list]
        result_alias_list = list(group_by_list)
        result_col_types = [{"name": gb, "type": col_type_map.get(gb, "")} for gb in group_by_list]
        used_aliases = set(group_by_list)

        for idx, agg in enumerate(aggs_list):
            op_raw = str(agg.get("op", "")).upper().strip()
            target_col = (agg.get("col") or "").strip()
            user_alias = (agg.get("alias") or "").strip()

            if op_raw not in _AGG_OPS:
                raise HTTPException(400, f"不支持的聚合操作: {op_raw}")

            if op_raw == "COUNT":
                if not target_col or target_col == "*":
                    expr = "COUNT(*)"
                    default_alias = "count"
                else:
                    if target_col not in col_names:
                        raise HTTPException(400, f"列 '{target_col}' 不存在")
                    expr = f'COUNT("{target_col}")'
                    default_alias = f"count_{target_col}"
            else:
                if not target_col or target_col == "*":
                    raise HTTPException(400, f"{op_raw} 必须指定目标列")
                if target_col not in col_names:
                    raise HTTPException(400, f"列 '{target_col}' 不存在")
                if op_raw in ("SUM", "AVG") and not _is_numeric_type(col_type_map.get(target_col, "")):
                    raise HTTPException(400, f"{op_raw} 只能用于数值列，'{target_col}' 类型是 {col_type_map.get(target_col)}")
                expr = f'{op_raw}("{target_col}")'
                default_alias = f"{op_raw.lower()}_{target_col}"

            alias = user_alias or default_alias
            # 避免别名冲突
            base = alias
            n = 2
            while alias in used_aliases:
                alias = f"{base}_{n}"
                n += 1
            used_aliases.add(alias)

            select_parts.append(f'{expr} AS "{alias}"')
            result_alias_list.append(alias)
            result_col_types.append({"name": alias, "type": "DOUBLE" if op_raw == "AVG" else ("BIGINT" if op_raw == "COUNT" else col_type_map.get(target_col, ""))})

        # 复用 WHERE / SEARCH 构造
        params = []
        where_clause, params = build_where(json.loads(filters) if filters else [], params, col_list)
        if search:
            search_clause, params = build_search_clause(col_list, search, params)
            if search_clause:
                where_clause = (where_clause + " AND " + search_clause) if where_clause else ("WHERE " + search_clause)

        group_by_sql = ", ".join(f'"{gb}"' for gb in group_by_list)

        # 排序：默认按第一个聚合项 DESC，否则按第一个 group_by ASC
        if sort_col and sort_col in used_aliases:
            order_alias = sort_col
        elif len(result_alias_list) > len(group_by_list):
            order_alias = result_alias_list[len(group_by_list)]
            sort_dir = sort_dir or "desc"
        else:
            order_alias = group_by_list[0]
            sort_dir = sort_dir or "asc"
        order_direction = "DESC" if sort_dir.lower() == "desc" else "ASC"

        offset = (page - 1) * page_size

        # 翻页时跳过 distinct group count 全表扫描
        total = None
        if not skip_count:
            total = conn.execute(
                f'SELECT COUNT(*) FROM (SELECT {group_by_sql} FROM "{table_name}" {where_clause} GROUP BY {group_by_sql}) AS _grp',
                params,
            ).fetchone()[0]

        select_sql = ", ".join(select_parts)
        sql = (
            f'SELECT {select_sql} FROM "{table_name}" {where_clause} '
            f'GROUP BY {group_by_sql} ORDER BY "{order_alias}" {order_direction} NULLS LAST '
            f'LIMIT {page_size} OFFSET {offset}'
        )
        rows_raw = conn.execute(sql, params).fetchall()
        rows = [dict(zip(result_alias_list, row)) for row in rows_raw]
        for row in rows:
            for k, v in row.items():
                if not isinstance(v, (str, int, float, bool, type(None))):
                    row[k] = str(v)

        return {
            "columns": result_col_types,
            "rows": rows,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (max(1, (total + page_size - 1) // page_size) if total is not None else None),
            "group_by": group_by_list,
            "aggs": aggs_list,
            "skipped_count": skip_count,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


# ===== DELETE TABLE =====
@router.delete("/table/{table_name}")
async def delete_table(table_name: str, user=Depends(require_permission("delete"))):
    conn = get_user_db(user["sub"])
    try:
        conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
        return {"success": True}
    except Exception as e:
        raise HTTPException(500, str(e))


# ===== CLEAR TABLE DATA =====
@router.delete("/table/{table_name}/data")
async def clear_table_data(
    table_name: str,
    filters: Optional[str] = None,
    search: Optional[str] = None,
    user=Depends(require_permission("delete")),
):
    conn = get_user_db(user["sub"])
    try:
        params = []
        cols = conn.execute(f'DESCRIBE "{table_name}"').fetchall()
        col_list = [{"name": c[0], "type": c[1]} for c in cols]
        where_clause, params = build_where(json.loads(filters) if filters else [], params, col_list)

        if search:
            search_clause, params = build_search_clause(col_list, search, params)
            if search_clause:
                if where_clause:
                    where_clause = where_clause + " AND " + search_clause
                else:
                    where_clause = "WHERE " + search_clause

        count = conn.execute(f'SELECT COUNT(*) FROM "{table_name}" {where_clause}', params).fetchone()[0]
        conn.execute(f'DELETE FROM "{table_name}" {where_clause}', params)
        return {"success": True, "table": table_name, "deleted_count": count}
    except Exception as e:
        raise HTTPException(500, str(e))


# ===== FILE MANAGEMENT =====
@router.get("/files")
async def list_files(user=Depends(require_permission("upload"))):
    user_id = user["sub"]
    storage = get_storage()
    files = storage.list_files(user_id)
    result = []
    for f in files:
        path = storage.get_local_path(user_id, f)
        size = path.stat().st_size if path.exists() else 0
        result.append({"name": f, "size": size})
    return {"files": result}


@router.delete("/files/{filename:path}")
async def delete_file(filename: str, user=Depends(require_permission("upload"))):
    user_id = user["sub"]
    storage = get_storage()
    try:
        storage.delete(user_id, filename)
        return {"success": True}
    except Exception as e:
        raise HTTPException(500, str(e))


@router.delete("/files")
async def clear_all_files(user=Depends(require_permission("upload"))):
    user_id = user["sub"]
    storage = get_storage()
    files = storage.list_files(user_id)
    for f in files:
        storage.delete(user_id, f)
    return {"success": True, "deleted": len(files)}
